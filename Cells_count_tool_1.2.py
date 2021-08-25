# -*- coding: utf-8 -*-
"""
Created on Thu Aug  5 11:45:11 2021

@author: ar3134ba
"""

import sys
from PyQt5.QtWidgets import QApplication, QInputDialog, QWidget, QPushButton, QVBoxLayout, QLabel, QFileDialog
from PyQt5.QtGui import QIcon, QPixmap, QFont
from PyQt5.QtCore import Qt
import cv2
import os, glob
import datetime
import pandas as pd
import numpy as np
from openpyxl import load_workbook

class Cells_counts():   
            
    def cells_detection(self, file, threshold):
        print('analyze start')
    
        '''This function create a gray picture with the jpg file. 
            Then, with the cv2 library we can count the cells '''
        
        
        # Read image
        img = cv2.imread(file)
        
        # Convert image to grayscale
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        
        #NEW 2021-08-16 Threshold
        thresholdValue = threshold
        maxValPixel = 255
        
        #Apply threshold
        ret, thresh2 = cv2.threshold(gray, thresholdValue, maxValPixel, cv2.THRESH_BINARY)
        
        ##Value pixel RGB
        # channels  = cv2.mean(thresh2)
        # observation = np.array([(channels[2], channels[1], channels[0])])
        # print('HERREEEEEEEEEEE*',observation)
        
        # cv2.imshow('Binary Threshold Inverted', thresh2)
        
        # Size of the picture
        width,height,dimension = img.shape
          # half of the original image
        required_width, required_height = width // 2, height // 2
            # resizing of an image is done
        resize_img_1 = cv2.resize(img,(required_height, required_width))
        
        # Show grayscale image
        #cv2.imshow('gray image', resize_img)
        cv2.waitKey(0)
        
        # Find contours
        contours, h = cv2.findContours(thresh2, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
        
        area = cv2.contourArea(contours[0])
        print('ZUUKOU', area)
        
        #Draw contours on the original picture
        image = cv2.drawContours(img, contours, -1, (0, 255, 0), 2)
        
        # Size of the picture
        width,height,dimension = img.shape
          # half of the original image
        required_width, required_height = width // 2, height // 2
            # resizing of an image is done
        resize_img_2 = cv2.resize(image,(required_height, required_width))
        
        # Show edge imqge
        cv2.imshow('Original picture', resize_img_1)
        #cv2.imshow('Python contours', resize_img_2)
        cv2.imshow("contour_detected_image", resize_img_2)
        
        #Location for save the data
        parent_dir = os.getcwd()
        directory = 'Results_pictures'
        path = os.path.join(parent_dir, directory)
        
        if os.path.isdir(path):
            next
        else : 
            os.mkdir(path)  #New directory for results
            
        os.chdir(path)
        filename = file.split('/')
        filename_threshold = filename[len(filename)-1][:-4] + '_THRESHOLD_'+str(thresholdValue)+'_python.jpg'
        filename = filename[len(filename)-1][:-4] + '_python.jpg'
        
        #Save the python picture
        #cv2.imwrite(filename,resize_img_2)
        cv2.imwrite(filename_threshold,thresh2)
        cv2.waitKey(0)
        os.chdir(parent_dir)
        
        nbr_cells =len(contours)
        self.nbr_cells = nbr_cells
        self.filename_threshold = filename_threshold
        print(" ")
        print("Picture :", file)
        print('Number of cells found :', nbr_cells)  
        print(" ")
        lst_results = [filename, nbr_cells, threshold]
        
        return lst_results
    
    def save_xlsx(self, threshold):
        
        parent_dir = os.getcwd()
        xlsx_file = "PYCELL.xlsx"
        path_xlsx = os.path.join(parent_dir, xlsx_file)
        date = datetime.datetime.now()
        format = "%Y-%m-%d %H:%M:%S"
        date = date.strftime(format)
        
        d_cells = {'Date' : [date], 'Filename':[self.filename_threshold], 'Number of cells':[self.nbr_cells], 'Threshold value':[threshold]}
        df1 = pd.DataFrame(d_cells)
        # writer = pd.ExcelWriter('PYCELL.xlsx')
        
        if os.path.isfile(path_xlsx):

            self.append_df_to_excel(path_xlsx, df1,sheet_name='Results', header=None, index=False)
            
        else:
            
            df1.to_excel('PYCELL.xlsx', sheet_name='Results',index=False)

        
    def append_df_to_excel(self,filename, df, sheet_name='Sheet1', startrow=None,
                   truncate_sheet=False, 
                   **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.
    
        @param filename: File path or existing ExcelWriter
                         (Example: '/path/to/file.xlsx')
        @param df: DataFrame to save to workbook
        @param sheet_name: Name of sheet which will contain DataFrame.
                           (default: 'Sheet1')
        @param startrow: upper left cell row to dump data frame.
                         Per default (startrow=None) calculate the last row
                         in the existing DF and write to the next row...
        @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                               before writing DataFrame to Excel file
        @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                                [can be a dictionary]
        @return: None
    
        (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
        """
        # Excel file doesn't exist - saving and exiting
        if not os.path.isfile(filename):
            df.to_excel(
                filename,
                sheet_name=sheet_name, 
                startrow=startrow if startrow is not None else 0, 
                **to_excel_kwargs)
            return
        
        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')
    
        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
    
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
    
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    
        if startrow is None:
            startrow = 0
    
        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    
        # save the workbook
        writer.save()
            
class HomePage(QWidget):
    '''
    First page managing
    '''
    
    def __init__(self):
        super().__init__()
        self.title='Cells count tool'
        self.setWindowIcon(QIcon('icon_cell.png'))
        self.left=10
        self.top=10
        width=300 #300
        height=400 #400
        self.setFixedSize(width, height)
        self.initUI()
        
       
    def initUI(self):
        
        self.setWindowTitle(self.title)
        self.setWindowIcon(QIcon('icon_cell.png'))
        
        layout = QVBoxLayout()
        label=QLabel()
        pixmap = QPixmap('lund_logo.png')
        pixmap = pixmap.scaled(250,250)
        label.setPixmap(pixmap)
        label.setAlignment(Qt.AlignCenter)

        label2 = QLabel('Welcome on the cells counts tool !',self)
        label2.setFont(QFont('Arial', 13))
        label2.setAlignment(Qt.AlignCenter)

        button=QPushButton('Load a file',self)
        button.clicked.connect(self.openFileNameDialog)
        button.setToolTip('.jpg / .png / .tiff')
        button.setFont(QFont('Arial', 13))
        
        # button2=QPushButton('Load multiple files',self)
        # button2.setToolTip('JPG file')
        # button2.setFont(QFont('Arial', 13))
        
        layout.addWidget(label)
        layout.addWidget(label2)
        layout.addWidget(button, alignment=Qt.AlignCenter)
        # layout.addWidget(button2, alignment=Qt.AlignCenter)
        
        self.setLayout(layout)
        self.show()
    
    def openFileNameDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog

        global filename

        filename, _ = QFileDialog.getOpenFileName(self,"File Explorer", "","All Files (*);;Python Files (*.py)", options=options)
        if filename:
            print(filename)
        
        global nbr_cells
        global name_picture
        global threshold
        
        threshold = 20
        call_method=Cells_counts()
        lst = call_method.cells_detection(filename, threshold)
        call_method.save_xlsx(threshold)
        name_picture = lst[0]
        nbr_cells = lst[1]
        
        self.name_picture = name_picture
        self.next=Second_Windows()
    
    def for_n_file(self):
        os.chdir('\\\\fysfile.nano.lu.se\\fyshome$\\H2$\\ar3134ba\\My Documents\\Python Scripts\\Nano_Lund\\jpg_nd2_files\\20200709_exp16Sav_A549_Ki67_Hoechst')
    
        for file in glob.glob("*.jpg"):
            Cells_counts.cells_detection(file, threshold)
    
class Second_Windows(QWidget):
    
    def __init__(self):
        super().__init__()
        self.title='Cells counts tool - Result'
        self.left=10
        self.top=10
        self.width=800
        self.height=240
        self.initUI()
        
    def display(self):
        integer , pressed = QInputDialog.getInt(self, "Threshold value","Number:",
                                            20, 0, 255, 1)
        if pressed:
            self.valuechange(integer)
            
         
    def initUI(self):
        self.setWindowIcon(QIcon('icon_cell.png'))
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)
        
        layout = QVBoxLayout()
        
        self.label1=QLabel("Picture : "+ str(name_picture), self)
        self.label1.setAlignment(Qt.AlignLeft)
        
        self.label2=QLabel("Save directory : "+ str(os.getcwd()+'\\Results_pictures'), self)
        self.label2.setAlignment(Qt.AlignLeft)
        
        self.label3=QLabel("Number of cells found : "+ str(nbr_cells), self)
        self.label3.setAlignment(Qt.AlignLeft)
        
        layout.addWidget(self.label1)
        layout.addWidget(self.label2)
        layout.addWidget(self.label3)
        
        self.label4 = QLabel('Threshold value = 20', self)
        self.label4.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        self.label4.setMinimumWidth(80)
        
        layout.addWidget(self.label4)
        
        self.label5 = QPushButton(self)
        self.label5.move(60,120)
        self.label5.setText('Update threshold')
        self.label5.clicked.connect(self.display)
        
        layout.addWidget(self.label5)
        
        self.setLayout(layout)
        self.show()
        
    def valuechange(self, threshold_select):
        
        self.label4.setText('Threshold value = ' +str(threshold_select))
        threshold = threshold_select
        call_method=Cells_counts()
        lst = call_method.cells_detection(filename, threshold)
        call_method.save_xlsx(threshold)
        self.label1.setText("Picture : "+ str(lst[0]))
        self.label3.setText("Number of cells found : "+ str(lst[1]))
        
    
if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex=HomePage()
    sys.exit(app.exec_())

