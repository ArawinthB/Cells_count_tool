# -*- coding: utf-8 -*-
"""
Created on Tue Aug  3 13:54:03 2021

@author: ar3134ba
"""
import datetime
import pandas as pd
import cv2
import os, glob
from openpyxl import load_workbook

        
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                   truncate_sheet=False, 
                   **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.
    
        @param filename: File path or existing ExcelWriter
                         (Example: '/path/to/file.xlsx')
        @param df: DataFrame to save to workbook
        @param sheet_name: Name of sheet which will contain DataFrame.
                           (default: 'Sheet1')
        @param startrow: upper left cell row to dump data frame.
                         Per default (startrow=None) calculate the last row
                         in the existing DF and write to the next row...
        @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                               before writing DataFrame to Excel file
        @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                                [can be a dictionary]
        @return: None
    
        (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
        """
        # Excel file doesn't exist - saving and exiting
        if not os.path.isfile(filename):
            df.to_excel(
                filename,
                sheet_name=sheet_name, 
                startrow=startrow if startrow is not None else 0, 
                **to_excel_kwargs)
            return
        
        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')
    
        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
    
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
    
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    
        if startrow is None:
            startrow = 0
    
        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    
        # save the workbook
        writer.save()
        

def save_xlsx(results):
    
    print(results)
    print(type(results))
    
    filename_threshold = results[0]
    nbr_cells = results[1]
    threshold_value = results[2]
    parent_dir = os.getcwd()
    xlsx_file = "PYCELL.xlsx"
    path_xlsx = os.path.join(parent_dir, xlsx_file)
    date = datetime.datetime.now()
    format = "%Y-%m-%d %H:%M:%S"
    date = date.strftime(format)
    
    d_cells = {'Date' : [date], 'Filename':[filename_threshold], 'Number of cells':[nbr_cells], 'Threshold value':[threshold_value]}
    df1 = pd.DataFrame(d_cells)
    # writer = pd.ExcelWriter('PYCELL.xlsx')
    
    if os.path.isfile(path_xlsx):
    
        append_df_to_excel(path_xlsx, df1,sheet_name='Results', header=None, index=False)
        
    else:
        
        df1.to_excel('PYCELL.xlsx', sheet_name='Results',index=False)
        
def cells_count(file, threshold):
    
    '''This function create a gray picture with the jpg file. 
        Then, with the cv2 library we can count the cells '''
        
    # Read image
    img = cv2.imread(file)
    
    # Convert image to grayscale
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    # Size of the picture
    width,height,dimension = img.shape
      # half of the original image
    required_width, required_height = width // 2, height // 2
        # resizing of an image is done
    resize_img = cv2.resize(gray,(required_height, required_width))
    
    # Show grayscale image
    #cv2.imshow('gray image', resize_img)
    cv2.waitKey(0)
       
    #NEW 2021-08-16 Threshold
    thresholdValue = threshold
    maxValPixel = 255
    
    ret, thresh2 = cv2.threshold(gray, thresholdValue, maxValPixel, cv2.THRESH_BINARY)
    
    # Find contours
    contours, h = cv2.findContours(thresh2, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
    
    # Size of the picture
    width,height,dimension = img.shape
      # half of the original image
    required_width, required_height = width // 2, height // 2
        # resizing of an image is done
    resize_img = cv2.resize(thresh2,(required_height, required_width))
    
    # Show edge imqge
    #cv2.imshow(file, resize_img)
    
    #Location for save the data
    directory_name ='Results_pictures'
    parent_dir = os.getcwd()

    filename = file.split('.')
    filename = filename[0] + '_python_Threshold_' + str(thresholdValue) +'.' + filename[1]  
    path = os.path.join(parent_dir, directory_name)
        
    if os.path.isdir(path):
        next
    else :
        os.mkdir(path)  #New directory for results
    
    #Save the python picture
    os.chdir(path)
    cv2.imwrite(filename,resize_img)
    cv2.waitKey(0)
    os.chdir(parent_dir)
    
    nbr_cells = len(contours)
    
    print(" ")
    print("Picture :", file)
    print('Number of cells found :',nbr_cells )  
    print(" ")
    
    
    return filename, nbr_cells, thresholdValue

if __name__ == '__main__':
    
    # Location of the files
    #os.chdir('\\\\fysfile.nano.lu.se\\fyshome$\\H2$\\ar3134ba\\My Documents\\Python Scripts\\Nano_Lund\\jpg_nd2_files\\20200709_exp16Sav_A549_Ki67_Hoechst')
    os.getcwd()
    
    for file in glob.glob("*.jpg"):
        print(file)
        threshold = 20
        results = cells_count(file, threshold)
        save_xlsx(results)
        
        


